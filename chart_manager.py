"""
Chart manager for IP SLA data visualization.

Creates and manages Excel charts with predefined favorites.
"""

import logging
from datetime import datetime
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from .config import COLUMNS, CHART_FAVORITES, DATA_SHEET, CHART_SHEET_PREFIX

logger = logging.getLogger(__name__)


class ChartManager:
    """
    Manages chart creation for IP SLA measurement data.
    
    Supports predefined favorite charts and custom chart generation
    with time range filtering.
    """

    # Chart styling constants
    CHART_WIDTH = 18
    CHART_HEIGHT = 10
    
    # Color palette for chart series
    COLORS = [
        "4472C4",  # Blue
        "ED7D31",  # Orange
        "A5A5A5",  # Gray
        "FFC000",  # Gold
        "5B9BD5",  # Light blue
        "70AD47",  # Green
        "9E480E",  # Brown
        "997300",  # Olive
    ]

    def __init__(self, workbook: Workbook):
        """
        Initialize chart manager.
        
        Args:
            workbook: openpyxl Workbook object
        """
        self.workbook = workbook
        self.data_sheet = workbook[DATA_SHEET] if DATA_SHEET in workbook.sheetnames else None

    def get_column_index(self, column_name: str) -> Optional[int]:
        """
        Get 1-based column index for a column name.
        
        Args:
            column_name: Name of the column
            
        Returns:
            1-based column index or None if not found
        """
        try:
            return COLUMNS.index(column_name) + 1
        except ValueError:
            logger.warning(f"Column not found: {column_name}")
            return None

    def get_row_range_for_dates(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Tuple[int, int]:
        """
        Find row indices for a date range.
        
        Args:
            start_date: Start of range (inclusive), None for beginning
            end_date: End of range (inclusive), None for end
            
        Returns:
            Tuple of (first_row, last_row) - 1-based indices
        """
        if self.data_sheet is None:
            return 2, 2
        
        first_row = None
        last_row = 2
        
        for row_idx in range(2, self.data_sheet.max_row + 1):
            cell_value = self.data_sheet.cell(row=row_idx, column=1).value
            if cell_value is None:
                continue
            
            # Convert to datetime if string
            if isinstance(cell_value, str):
                try:
                    cell_value = datetime.strptime(cell_value, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue
            
            # Check if within range
            if start_date and cell_value < start_date:
                continue
            if end_date and cell_value > end_date:
                continue
            
            if first_row is None:
                first_row = row_idx
            last_row = row_idx
        
        return first_row or 2, last_row

    def create_chart(
        self,
        chart_name: str,
        title: str,
        y_columns: List[str],
        y_label: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        sheet_name: Optional[str] = None
    ) -> Optional[str]:
        """
        Create a line chart in the workbook.
        
        Args:
            chart_name: Identifier for the chart
            title: Chart title
            y_columns: List of column names for Y-axis data
            y_label: Label for Y-axis
            start_date: Start of date range (optional)
            end_date: End of date range (optional)
            sheet_name: Name for the chart sheet (optional)
            
        Returns:
            Name of the created sheet or None if failed
        """
        if self.data_sheet is None:
            logger.error("No data sheet available")
            return None
        
        # Determine row range
        first_row, last_row = self.get_row_range_for_dates(start_date, end_date)
        
        if first_row >= last_row:
            logger.warning(f"No data in date range for chart: {chart_name}")
            return None
        
        # Create chart sheet
        target_sheet_name = sheet_name or f"{CHART_SHEET_PREFIX}{chart_name}"
        
        # Remove existing sheet with same name
        if target_sheet_name in self.workbook.sheetnames:
            del self.workbook[target_sheet_name]
        
        chart_sheet = self.workbook.create_sheet(target_sheet_name)
        
        # Create line chart
        chart = LineChart()
        chart.title = title
        chart.style = 10
        chart.y_axis.title = y_label
        chart.x_axis.title = "Time"
        chart.width = self.CHART_WIDTH
        chart.height = self.CHART_HEIGHT
        
        # Configure x-axis for dates
        chart.x_axis.number_format = 'MM/DD HH:MM'
        chart.x_axis.majorTimeUnit = "days"
        
        # Add time reference (X-axis)
        time_col = self.get_column_index("StartTime")
        dates = Reference(
            self.data_sheet,
            min_col=time_col,
            min_row=first_row,
            max_row=last_row
        )
        
        # Add data series
        for idx, col_name in enumerate(y_columns):
            col_idx = self.get_column_index(col_name)
            if col_idx is None:
                continue
            
            data = Reference(
                self.data_sheet,
                min_col=col_idx,
                min_row=first_row - 1,  # Include header for series name
                max_row=last_row
            )
            
            chart.add_data(data, titles_from_data=True)
        
        chart.set_categories(dates)
        
        # Apply colors to series
        for idx, series in enumerate(chart.series):
            series.graphicalProperties.line.solidFill = self.COLORS[idx % len(self.COLORS)]
            series.graphicalProperties.line.width = 20000  # EMUs
            series.smooth = False
        
        # Add chart to sheet
        chart_sheet.add_chart(chart, "A1")
        
        # Add metadata
        chart_sheet.cell(row=1, column=15, value="Chart Info")
        chart_sheet.cell(row=2, column=15, value="Start Date:")
        chart_sheet.cell(row=2, column=16, value=start_date.isoformat() if start_date else "All")
        chart_sheet.cell(row=3, column=15, value="End Date:")
        chart_sheet.cell(row=3, column=16, value=end_date.isoformat() if end_date else "All")
        chart_sheet.cell(row=4, column=15, value="Columns:")
        chart_sheet.cell(row=4, column=16, value=", ".join(y_columns))
        
        logger.info(f"Created chart: {target_sheet_name}")
        return target_sheet_name

    def create_favorite_chart(
        self,
        favorite_name: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Optional[str]:
        """
        Create a predefined favorite chart.
        
        Args:
            favorite_name: Name of the favorite (from CHART_FAVORITES)
            start_date: Start of date range (optional)
            end_date: End of date range (optional)
            
        Returns:
            Name of the created sheet or None if failed
        """
        if favorite_name not in CHART_FAVORITES:
            logger.error(f"Unknown favorite: {favorite_name}")
            return None
        
        fav = CHART_FAVORITES[favorite_name]
        
        return self.create_chart(
            chart_name=favorite_name,
            title=fav["title"],
            y_columns=fav["y_columns"],
            y_label=fav["y_label"],
            start_date=start_date,
            end_date=end_date
        )

    def create_all_favorite_charts(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> List[str]:
        """
        Create all predefined favorite charts.
        
        Args:
            start_date: Start of date range (optional)
            end_date: End of date range (optional)
            
        Returns:
            List of created sheet names
        """
        created = []
        for name in CHART_FAVORITES:
            sheet_name = self.create_favorite_chart(name, start_date, end_date)
            if sheet_name:
                created.append(sheet_name)
        return created

    def list_favorites(self) -> List[Tuple[str, str]]:
        """
        Get list of available favorite charts.
        
        Returns:
            List of (name, title) tuples
        """
        return [(name, fav["title"]) for name, fav in CHART_FAVORITES.items()]
