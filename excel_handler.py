"""
Excel handler for IP SLA measurements.

Handles reading, writing, and deduplication of measurement data.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Set, Tuple, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .config import OUTPUT_FILE, DATA_SHEET, COLUMNS
from .parser import IPSLARecord

logger = logging.getLogger(__name__)


class ExcelHandler:
    """
    Manages Excel workbook operations for IP SLA data.
    
    Handles creation, reading, writing, and deduplication
    of measurement records in the Excel file.
    """

    # Styling constants
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, filepath: Path = OUTPUT_FILE):
        """
        Initialize Excel handler.
        
        Args:
            filepath: Path to the Excel file
        """
        self.filepath = filepath
        self.workbook: Optional[Workbook] = None
        self.data_sheet: Optional[Worksheet] = None

    def open_or_create(self) -> None:
        """Open existing workbook or create new one with headers."""
        if self.filepath.exists():
            logger.info(f"Opening existing workbook: {self.filepath}")
            self.workbook = load_workbook(self.filepath)
            
            # Get or create data sheet
            if DATA_SHEET in self.workbook.sheetnames:
                self.data_sheet = self.workbook[DATA_SHEET]
            else:
                self.data_sheet = self.workbook.create_sheet(DATA_SHEET, 0)
                self._write_headers()
        else:
            logger.info(f"Creating new workbook: {self.filepath}")
            self.workbook = Workbook()
            self.data_sheet = self.workbook.active
            self.data_sheet.title = DATA_SHEET
            self._write_headers()

    def _write_headers(self) -> None:
        """Write column headers with formatting."""
        for col_idx, header in enumerate(COLUMNS, start=1):
            cell = self.data_sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        
        # Set column widths
        column_widths = {
            'StartTime': 20,
            'MinMOS': 10, 'MaxMOS': 10,
            'MinICPIF': 10, 'MaxICPIF': 10,
            'NumRTT': 12,
            'RTT_Min_ms': 12, 'RTT_Avg_ms': 12, 'RTT_Max_ms': 12,
            'RTT_Over_Threshold_Count': 22, 'RTT_Over_Threshold_Pct': 20,
            'Num_OneWay_Samples': 18,
            'Jitter_SD_Avg_ms': 16, 'Jitter_SD_Max_ms': 16,
            'Jitter_DS_Avg_ms': 16, 'Jitter_DS_Max_ms': 16,
            'Loss_SD': 10, 'Loss_DS': 10,
            'Packet_Late_Arrival': 18, 'OutOfSeq': 10, 'TailDrop': 10,
            'Successes': 12, 'Failures': 10
        }
        
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            width = column_widths.get(col_name, 12)
            self.data_sheet.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        self.data_sheet.freeze_panes = 'A2'

    def get_existing_timestamps(self) -> Set[datetime]:
        """
        Get set of existing timestamps to prevent duplicates.
        
        Returns:
            Set of datetime objects for existing records
        """
        timestamps = set()
        
        if self.data_sheet is None:
            return timestamps
        
        # Skip header row
        for row in self.data_sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                # Handle both datetime objects and strings
                if isinstance(row[0], datetime):
                    timestamps.add(row[0])
                elif isinstance(row[0], str):
                    try:
                        ts = datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")
                        timestamps.add(ts)
                    except ValueError:
                        pass
        
        logger.debug(f"Found {len(timestamps)} existing timestamps")
        return timestamps

    def append_records(self, records: List[IPSLARecord]) -> Tuple[int, int]:
        """
        Append new records, skipping duplicates.
        
        Args:
            records: List of IPSLARecord objects to append
            
        Returns:
            Tuple of (added_count, skipped_count)
        """
        existing = self.get_existing_timestamps()
        added = 0
        skipped = 0
        
        # Find next empty row
        next_row = self.data_sheet.max_row + 1
        if next_row == 2 and self.data_sheet.cell(row=2, column=1).value is None:
            next_row = 2  # First data row if sheet is empty
        
        for record in records:
            if record.start_time in existing:
                logger.debug(f"Skipping duplicate: {record.start_time}")
                skipped += 1
                continue
            
            # Write record
            row_data = record.to_row()
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.data_sheet.cell(row=next_row, column=col_idx, value=value)
                cell.border = self.BORDER
                
                # Format datetime column
                if col_idx == 1 and isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                    cell.alignment = Alignment(horizontal='center')
            
            existing.add(record.start_time)
            next_row += 1
            added += 1
        
        logger.info(f"Added {added} records, skipped {skipped} duplicates")
        return added, skipped

    def sort_by_timestamp(self) -> None:
        """Sort all data rows by timestamp (ascending)."""
        if self.data_sheet is None or self.data_sheet.max_row <= 1:
            return
        
        # Collect all data rows
        data_rows = []
        for row in self.data_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                data_rows.append(list(row))
        
        # Sort by timestamp (first column)
        data_rows.sort(key=lambda x: x[0] if isinstance(x[0], datetime) else datetime.min)
        
        # Clear existing data (keep headers)
        for row_idx in range(2, self.data_sheet.max_row + 1):
            for col_idx in range(1, len(COLUMNS) + 1):
                self.data_sheet.cell(row=row_idx, column=col_idx).value = None
        
        # Rewrite sorted data
        for row_idx, row_data in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.data_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER
                if col_idx == 1 and isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                    cell.alignment = Alignment(horizontal='center')
        
        logger.info("Data sorted by timestamp")

    def get_data_range(self) -> Tuple[Optional[datetime], Optional[datetime]]:
        """
        Get the time range of existing data.
        
        Returns:
            Tuple of (earliest_timestamp, latest_timestamp) or (None, None)
        """
        timestamps = self.get_existing_timestamps()
        if not timestamps:
            return None, None
        return min(timestamps), max(timestamps)

    def get_all_data(self) -> List[List]:
        """
        Get all data rows from the sheet.
        
        Returns:
            List of row data (including header)
        """
        if self.data_sheet is None:
            return []
        
        data = []
        for row in self.data_sheet.iter_rows(values_only=True):
            data.append(list(row))
        return data

    def save(self) -> None:
        """Save the workbook to disk."""
        if self.workbook:
            self.workbook.save(self.filepath)
            logger.info(f"Workbook saved: {self.filepath}")

    def close(self) -> None:
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.data_sheet = None
